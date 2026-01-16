"""
Metadata extraction module for Airganizer Stage 1
"""

import os
import hashlib
import mimetypes
from pathlib import Path
from typing import Dict, Any, Optional
from datetime import datetime
import json


class FileMetadata:
    """Container for file metadata"""
    
    def __init__(self, file_path: Path):
        """
        Initialize file metadata
        
        Args:
            file_path: Path to the file
        """
        self.file_path = file_path
        self.metadata: Dict[str, Any] = {}
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert metadata to dictionary"""
        return self.metadata
    
    def to_json(self) -> str:
        """Convert metadata to JSON string"""
        return json.dumps(self.metadata, indent=2, default=str)


class MetadataExtractor:
    """Extract metadata from files"""
    
    def __init__(self, calculate_hash: bool = False):
        """
        Initialize metadata extractor
        
        Args:
            calculate_hash: Whether to calculate file hash
        """
        self.calculate_hash = calculate_hash
        
        # Initialize mimetypes
        mimetypes.init()
    
    def extract_basic_metadata(self, file_path: Path) -> Dict[str, Any]:
        """
        Extract basic file metadata (always available)
        
        Args:
            file_path: Path to the file
        
        Returns:
            Dictionary of basic metadata
        """
        stat = file_path.stat()
        
        metadata = {
            'file_name': file_path.name,
            'file_path': str(file_path.absolute()),
            'file_extension': file_path.suffix.lower(),
            'file_size': stat.st_size,
            'file_size_human': self._human_readable_size(stat.st_size),
            'created_time': datetime.fromtimestamp(stat.st_ctime).isoformat(),
            'modified_time': datetime.fromtimestamp(stat.st_mtime).isoformat(),
            'accessed_time': datetime.fromtimestamp(stat.st_atime).isoformat(),
        }
        
        return metadata
    
    def extract_mime_type(self, file_path: Path) -> Dict[str, Optional[str]]:
        """
        Extract MIME type information
        
        Args:
            file_path: Path to the file
        
        Returns:
            Dictionary with MIME type information
        """
        mime_type, encoding = mimetypes.guess_type(str(file_path))
        
        return {
            'mime_type': mime_type,
            'encoding': encoding,
            'media_type': mime_type.split('/')[0] if mime_type else None,
        }
    
    def calculate_file_hash(self, file_path: Path, algorithm: str = 'md5') -> str:
        """
        Calculate file hash
        
        Args:
            file_path: Path to the file
            algorithm: Hash algorithm (md5, sha1, sha256)
        
        Returns:
            Hex digest of the file hash
        """
        hash_func = hashlib.new(algorithm)
        
        with open(file_path, 'rb') as f:
            while chunk := f.read(8192):
                hash_func.update(chunk)
        
        return hash_func.hexdigest()
    
    def extract_exif_data(self, file_path: Path) -> Optional[Dict[str, Any]]:
        """
        Extract EXIF data from images
        
        Args:
            file_path: Path to the image file
        
        Returns:
            Dictionary of EXIF data or None
        """
        try:
            from PIL import Image
            from PIL.ExifTags import TAGS
            
            image = Image.open(file_path)
            exif_data = image.getexif()
            
            if not exif_data:
                return None
            
            exif = {}
            for tag_id, value in exif_data.items():
                tag = TAGS.get(tag_id, tag_id)
                
                # Convert bytes to string
                if isinstance(value, bytes):
                    try:
                        value = value.decode()
                    except:
                        value = str(value)
                
                exif[tag] = value
            
            # Add image dimensions
            exif['width'] = image.width
            exif['height'] = image.height
            exif['format'] = image.format
            exif['mode'] = image.mode
            
            return exif
        
        except ImportError:
            return {'error': 'PIL/Pillow not installed'}
        except Exception as e:
            return {'error': str(e)}
    
    def extract_audio_metadata(self, file_path: Path) -> Optional[Dict[str, Any]]:
        """
        Extract metadata from audio files
        
        Args:
            file_path: Path to the audio file
        
        Returns:
            Dictionary of audio metadata or None
        """
        try:
            from mutagen import File as MutagenFile
            
            audio = MutagenFile(file_path)
            
            if audio is None:
                return None
            
            metadata = {
                'duration': audio.info.length if hasattr(audio, 'info') else None,
                'bitrate': audio.info.bitrate if hasattr(audio, 'info') and hasattr(audio.info, 'bitrate') else None,
                'sample_rate': audio.info.sample_rate if hasattr(audio, 'info') and hasattr(audio.info, 'sample_rate') else None,
                'channels': audio.info.channels if hasattr(audio, 'info') and hasattr(audio.info, 'channels') else None,
            }
            
            # Extract tags
            if audio.tags:
                tags = {}
                for key, value in audio.tags.items():
                    # Convert list values to string
                    if isinstance(value, list):
                        value = ', '.join(str(v) for v in value)
                    tags[key] = str(value)
                metadata['tags'] = tags
            
            return metadata
        
        except ImportError:
            return {'error': 'mutagen not installed'}
        except Exception as e:
            return {'error': str(e)}
    
    def extract_video_metadata(self, file_path: Path) -> Optional[Dict[str, Any]]:
        """
        Extract metadata from video files
        
        Args:
            file_path: Path to the video file
        
        Returns:
            Dictionary of video metadata or None
        """
        try:
            from pymediainfo import MediaInfo
            
            media_info = MediaInfo.parse(file_path)
            
            metadata = {}
            
            for track in media_info.tracks:
                if track.track_type == 'General':
                    metadata['duration'] = track.duration
                    metadata['file_size'] = track.file_size
                    metadata['format'] = track.format
                
                elif track.track_type == 'Video':
                    metadata['video'] = {
                        'codec': track.codec,
                        'width': track.width,
                        'height': track.height,
                        'frame_rate': track.frame_rate,
                        'bit_rate': track.bit_rate,
                    }
                
                elif track.track_type == 'Audio':
                    metadata['audio'] = {
                        'codec': track.codec,
                        'channels': track.channel_s,
                        'sample_rate': track.sampling_rate,
                        'bit_rate': track.bit_rate,
                    }
            
            return metadata
        
        except ImportError:
            return {'error': 'pymediainfo not installed'}
        except Exception as e:
            return {'error': str(e)}
    
    def extract_document_metadata(self, file_path: Path) -> Optional[Dict[str, Any]]:
        """
        Extract metadata from document files (PDF, Office)
        
        Args:
            file_path: Path to the document file
        
        Returns:
            Dictionary of document metadata or None
        """
        extension = file_path.suffix.lower()
        
        # PDF metadata
        if extension == '.pdf':
            return self._extract_pdf_metadata(file_path)
        
        # Office documents
        elif extension in ['.docx', '.xlsx', '.pptx']:
            return self._extract_office_metadata(file_path)
        
        return None
    
    def _extract_pdf_metadata(self, file_path: Path) -> Optional[Dict[str, Any]]:
        """Extract metadata from PDF files"""
        try:
            from PyPDF2 import PdfReader
            
            reader = PdfReader(file_path)
            
            metadata = {
                'num_pages': len(reader.pages),
                'info': {}
            }
            
            if reader.metadata:
                for key, value in reader.metadata.items():
                    # Remove leading slash from keys
                    clean_key = key[1:] if key.startswith('/') else key
                    metadata['info'][clean_key] = str(value)
            
            return metadata
        
        except ImportError:
            return {'error': 'PyPDF2 not installed'}
        except Exception as e:
            return {'error': str(e)}
    
    def _extract_office_metadata(self, file_path: Path) -> Optional[Dict[str, Any]]:
        """Extract metadata from Office files"""
        try:
            from openpyxl import load_workbook
            from docx import Document
            from pptx import Presentation
            
            extension = file_path.suffix.lower()
            
            if extension == '.docx':
                doc = Document(file_path)
                core_props = doc.core_properties
                
                return {
                    'author': core_props.author,
                    'created': core_props.created.isoformat() if core_props.created else None,
                    'modified': core_props.modified.isoformat() if core_props.modified else None,
                    'title': core_props.title,
                    'subject': core_props.subject,
                    'num_paragraphs': len(doc.paragraphs),
                }
            
            elif extension == '.xlsx':
                wb = load_workbook(file_path, read_only=True, data_only=True)
                props = wb.properties
                
                return {
                    'author': props.creator,
                    'created': props.created.isoformat() if props.created else None,
                    'modified': props.modified.isoformat() if props.modified else None,
                    'title': props.title,
                    'num_sheets': len(wb.sheetnames),
                    'sheet_names': wb.sheetnames,
                }
            
            elif extension == '.pptx':
                prs = Presentation(file_path)
                core_props = prs.core_properties
                
                return {
                    'author': core_props.author,
                    'created': core_props.created.isoformat() if core_props.created else None,
                    'modified': core_props.modified.isoformat() if core_props.modified else None,
                    'title': core_props.title,
                    'num_slides': len(prs.slides),
                }
        
        except ImportError:
            return {'error': 'Required library not installed (python-docx, openpyxl, or python-pptx)'}
        except Exception as e:
            return {'error': str(e)}
    
    def extract_all_metadata(self, file_path: Path, extract_exif: bool = True,
                            extract_audio: bool = True, extract_video: bool = True,
                            extract_document: bool = True) -> FileMetadata:
        """
        Extract all available metadata from a file
        
        Args:
            file_path: Path to the file
            extract_exif: Whether to extract EXIF data
            extract_audio: Whether to extract audio metadata
            extract_video: Whether to extract video metadata
            extract_document: Whether to extract document metadata
        
        Returns:
            FileMetadata object with all extracted metadata
        """
        metadata = FileMetadata(file_path)
        
        # Basic metadata (always extracted)
        metadata.metadata.update(self.extract_basic_metadata(file_path))
        
        # MIME type
        metadata.metadata.update(self.extract_mime_type(file_path))
        
        # File hash (optional)
        if self.calculate_hash:
            metadata.metadata['md5_hash'] = self.calculate_file_hash(file_path)
        
        # Media type specific metadata
        mime_type = metadata.metadata.get('mime_type', '')
        media_type = metadata.metadata.get('media_type', '')
        
        # EXIF data for images
        if extract_exif and media_type == 'image':
            exif_data = self.extract_exif_data(file_path)
            if exif_data:
                metadata.metadata['exif'] = exif_data
        
        # Audio metadata
        if extract_audio and media_type == 'audio':
            audio_data = self.extract_audio_metadata(file_path)
            if audio_data:
                metadata.metadata['audio_metadata'] = audio_data
        
        # Video metadata
        if extract_video and media_type == 'video':
            video_data = self.extract_video_metadata(file_path)
            if video_data:
                metadata.metadata['video_metadata'] = video_data
        
        # Document metadata
        if extract_document and mime_type in ['application/pdf', 
                                               'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                                               'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                               'application/vnd.openxmlformats-officedocument.presentationml.presentation']:
            doc_data = self.extract_document_metadata(file_path)
            if doc_data:
                metadata.metadata['document_metadata'] = doc_data
        
        return metadata
    
    @staticmethod
    def _human_readable_size(size_bytes: int) -> str:
        """
        Convert bytes to human readable size
        
        Args:
            size_bytes: Size in bytes
        
        Returns:
            Human readable size string
        """
        for unit in ['B', 'KB', 'MB', 'GB', 'TB']:
            if size_bytes < 1024.0:
                return f"{size_bytes:.2f} {unit}"
            size_bytes /= 1024.0
        
        return f"{size_bytes:.2f} PB"
